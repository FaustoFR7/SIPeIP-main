import csv
import json
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO, StringIO
from xml.sax.saxutils import escape

from django.core.serializers.json import DjangoJSONEncoder
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table as ExcelTable, TableStyleInfo
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .datasets import valor_json


MIME_TYPES = {
    "json": "application/json; charset=utf-8",
    "csv": "text/csv; charset=utf-8",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}


def _valor_tabular(valor):
    convertido = valor_json(valor)
    if convertido is None:
        return ""
    if isinstance(convertido, bool):
        return "Sí" if convertido else "No"
    if isinstance(convertido, (dict, list)):
        return json.dumps(convertido, ensure_ascii=False, cls=DjangoJSONEncoder)
    texto = str(convertido)
    # Evita que una hoja de cálculo interprete contenido controlado por usuarios.
    if texto.startswith(("=", "+", "-", "@")):
        return f"'{texto}"
    return texto


def _valor_xlsx(valor):
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        if timezone.is_aware(valor):
            return timezone.localtime(valor).replace(tzinfo=None)
        return valor
    if isinstance(valor, (date, int, float, Decimal, bool)):
        return valor
    if isinstance(valor, (dict, list)):
        return json.dumps(valor, ensure_ascii=False, cls=DjangoJSONEncoder)
    texto = str(valor)
    if texto.startswith(("=", "+", "-", "@")):
        return f"'{texto}"
    return texto


def exportar_json(spec, filas, metadata):
    contenido = {
        **metadata,
        "columnas": [
            {"campo": campo, "titulo": titulo} for campo, titulo in spec.columnas
        ],
        "resultados": [
            {clave: valor_json(valor) for clave, valor in fila.items()}
            for fila in filas
        ],
    }
    return json.dumps(
        contenido,
        ensure_ascii=False,
        indent=2,
        cls=DjangoJSONEncoder,
    ).encode("utf-8")


def exportar_csv(spec, filas, metadata):
    salida = StringIO(newline="")
    # BOM UTF-8 para apertura correcta en Excel sin alterar el CSV lógico.
    salida.write("\ufeff")
    escritor = csv.writer(salida)
    escritor.writerow([titulo for _, titulo in spec.columnas])
    for fila in filas:
        escritor.writerow([_valor_tabular(fila.get(campo)) for campo, _ in spec.columnas])
    return salida.getvalue().encode("utf-8")


def exportar_xlsx(spec, filas, metadata):
    libro = Workbook(write_only=False)
    hoja = libro.active
    hoja.title = "Reporte"
    hoja.sheet_view.showGridLines = False
    hoja.freeze_panes = "A2"
    hoja.append([titulo for _, titulo in spec.columnas])
    relleno = PatternFill("solid", fgColor="17365D")
    for celda in hoja[1]:
        celda.font = Font(color="FFFFFF", bold=True)
        celda.fill = relleno

    for fila in filas:
        hoja.append([_valor_xlsx(fila.get(campo)) for campo, _ in spec.columnas])

    for fila in hoja.iter_rows(min_row=2):
        for celda in fila:
            celda.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(celda.value, (datetime, date)):
                celda.number_format = "yyyy-mm-dd"
            elif isinstance(celda.value, (Decimal, float)):
                celda.number_format = "#,##0.00"

    for indice, (_, titulo) in enumerate(spec.columnas, start=1):
        valores = [
            len(str(hoja.cell(row=fila, column=indice).value or ""))
            for fila in range(1, min(hoja.max_row, 100) + 1)
        ]
        ancho = min(max(max(valores, default=len(titulo)) + 2, 12), 36)
        hoja.column_dimensions[get_column_letter(indice)].width = ancho
    hoja.auto_filter.ref = hoja.dimensions
    if hoja.max_row > 1:
        tabla = ExcelTable(displayName="ReporteSIPeIP", ref=hoja.dimensions)
        tabla.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        hoja.add_table(tabla)

    informacion = libro.create_sheet("Información")
    informacion.sheet_view.showGridLines = False
    informacion.append(["Reporte", spec.nombre])
    informacion.append(["Generado", metadata["generado_en"]])
    informacion.append(["Registros", metadata["total"]])
    informacion.append(
        [
            "Filtros",
            json.dumps(metadata.get("filtros", {}), ensure_ascii=False),
        ]
    )
    informacion.column_dimensions["A"].width = 18
    informacion.column_dimensions["B"].width = 72
    for celda in informacion["A"]:
        celda.font = Font(bold=True, color="17365D")
    for fila in informacion.iter_rows():
        for celda in fila:
            celda.alignment = Alignment(vertical="top", wrap_text=True)

    salida = BytesIO()
    libro.save(salida)
    return salida.getvalue()


def exportar_pdf(spec, filas, metadata):
    salida = BytesIO()
    documento = SimpleDocTemplate(
        salida,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=spec.nombre,
        author="SIPeIP",
    )
    estilos = getSampleStyleSheet()
    titulo = ParagraphStyle(
        "TituloReporte",
        parent=estilos["Title"],
        fontName="Helvetica-Bold",
        fontSize=15,
        leading=18,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#17365D"),
    )
    celda = ParagraphStyle(
        "CeldaReporte",
        parent=estilos["BodyText"],
        fontName="Helvetica",
        fontSize=6.5,
        leading=8,
    )
    encabezado = ParagraphStyle(
        "EncabezadoReporte",
        parent=celda,
        fontName="Helvetica-Bold",
        textColor=colors.white,
    )
    elementos = [
        Paragraph(spec.nombre, titulo),
        Paragraph(
            f"Generado: {metadata['generado_en']} | Registros: {metadata['total']}",
            estilos["Normal"],
        ),
        Spacer(1, 4 * mm),
    ]
    datos = [[Paragraph(str(titulo_columna), encabezado) for _, titulo_columna in spec.columnas]]
    for fila in filas:
        datos.append(
            [
                Paragraph(
                    escape(_valor_tabular(fila.get(campo))[:500]),
                    celda,
                )
                for campo, _ in spec.columnas
            ]
        )

    ancho_util = landscape(A4)[0] - 20 * mm
    anchos = [ancho_util / max(len(spec.columnas), 1)] * len(spec.columnas)
    tabla = Table(datos, repeatRows=1, colWidths=anchos)
    tabla.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17365D")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B8C4CE")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F5F7")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )
    elementos.append(tabla)

    def dibujar_pie(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#52606D"))
        canvas.drawString(10 * mm, 6 * mm, "SIPeIP - Reporte institucional")
        canvas.drawRightString(
            landscape(A4)[0] - 10 * mm,
            6 * mm,
            f"Página {doc.page}",
        )
        canvas.restoreState()

    documento.build(
        elementos,
        onFirstPage=dibujar_pie,
        onLaterPages=dibujar_pie,
    )
    return salida.getvalue()


EXPORTADORES = {
    "json": exportar_json,
    "csv": exportar_csv,
    "xlsx": exportar_xlsx,
    "pdf": exportar_pdf,
}
